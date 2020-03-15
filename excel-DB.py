# TODO: Move to scripts

import openpyxl
from users.models import User
from news.models import News, Courses
from organization.models import Organization


def export_from_excel(file, add_func):
    wb = openpyxl.load_workbook(filename=file)
    sheet = wb.active
    for i in range(1, sheet.max_row + 1):
        params = [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1)]
        add_func(*params)


if __name__ == '__main__':
    export_from_excel('test_models/orgs.xlsx', Organization.new)
    export_from_excel('test_models/users.xlsx', User.new)
    export_from_excel('test_models/news.xlsx', News.new)
    export_from_excel('test_models/courses.xlsx', Courses.new)
