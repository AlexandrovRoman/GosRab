import openpyxl
from users.models import User, Course, T2Form
from news.models import News
from organization.models import Organization, Vacancy
from app.config import here
from os import path


def export_from_excel(file, add_func):
    wb = openpyxl.load_workbook(filename=file)
    sheet = wb.active
    for i in range(1, sheet.max_row + 1):
        params = [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1)]
        add_func(*params)


def _abs_xls_path(filename):
    filename = filename if "." in filename else filename + ".xlsx"
    return path.join(here, "test_models", filename)


if __name__ == '__main__':
    export_from_excel(_abs_xls_path("orgs"), Organization.new)
    export_from_excel(_abs_xls_path("users"), User.new)
    export_from_excel(_abs_xls_path("t2"), T2Form.new)
    export_from_excel(_abs_xls_path("vacancies"), Vacancy.new)
    export_from_excel(_abs_xls_path("news"), News.new)
    export_from_excel(_abs_xls_path("courses"), Course.new)
